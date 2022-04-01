import os
import sys
from datetime import datetime
import openpyxl
from PIL import Image
from openpyxl.styles import PatternFill
from skimage import data, img_as_float
from skimage.metrics import structural_similarity as ssim
from skimage.metrics import peak_signal_noise_ratio as psnr
import pandas as pd
import numpy as np



def is_image_file(filename):
    # check if the image ends with png
    IMG_EXTENSIONS = [
        '.jpg', '.JPG', '.jpeg', '.JPEG',
        '.png', '.PNG', '.ppm', '.PPM', '.bmp', '.BMP',
    ]
    return any(filename.endswith(extension) for extension in IMG_EXTENSIONS)


def get_col_widths(dataframe):
    # First we find the maximum length of the index column
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]


def evaluate(dir_bg, dir_res, data_id='DATA SET NAME', dir_save=None):
    color_mean = PatternFill(patternType='solid',
                             fgColor='7777C9')
    color_median = PatternFill(patternType='solid',
                               fgColor='4B4B9F')
    color_std = PatternFill(patternType='solid',
                            fgColor='70709A')
    color_best = PatternFill(patternType='solid',
                             fgColor='31A91F')
    color_worst = PatternFill(patternType='solid',
                              fgColor='D81F1F')
    color_five_best = PatternFill(patternType='solid',
                                  fgColor='8FBC89')
    color_five_worst = PatternFill(patternType='solid',
                                   fgColor='DE8E7E')

    NBR_MARKED = 5

    psnr_list = []
    ssim_list = []
    image_list_gt = []
    image_list_result = []

    date = str(datetime.now())
    if dir_save is None:
        dir_save = os.path.join(os.path.dirname(dir_bg) + '/metrics_results_{}_{}.xlsx'.format(data_id, date))

    for img_bg, img_res in zip(os.listdir(dir_bg), os.listdir(dir_res)):
        ground_truth_image_path = os.path.join(dir_bg, img_bg)
        result_image_path = os.path.join(dir_res, img_res)
        if is_image_file(ground_truth_image_path) and is_image_file(result_image_path):
            can_load = True
            try:
                ground_truth_img = Image.open(ground_truth_image_path)
                result_img = Image.open(result_image_path)
            except:
                print('Was not able to load ' + ground_truth_image_path + 'or' + result_image_path)
                can_load = False
                continue
            if can_load and ground_truth_img.size == (1920, 1080) and result_img.size == (1920, 1080):
                image_list_gt.append(ground_truth_image_path)
                image_list_result.append(result_image_path)

                ground_truth_img = img_as_float(ground_truth_img)
                result_img = img_as_float(result_img)

                ssim_val = ssim(ground_truth_img, result_img, multichannel=True)
                psnr_val = psnr(ground_truth_img, result_img)
                ssim_list.append(ssim_val)
                psnr_list.append(psnr_val)

    d = {'GT': image_list_gt, 'Filtered': image_list_result, 'SSIM': ssim_list, 'PSNR': psnr_list}
    df = pd.DataFrame(data=d)
    df.to_excel(dir_save,
                index=False, sheet_name='sheet1')

    # mean and std values
    ssim_mean = np.mean(ssim_list)
    psnr_mean = np.mean(psnr_list)
    ssim_median = np.median(ssim_list)
    psnr_median = np.median(psnr_list)
    ssim_std = np.std(ssim_list)
    psnr_std = np.std(psnr_list)

    wb = openpyxl.load_workbook(dir_save)
    ws = wb['sheet1']  # Name of the working sheet

    max_ssim_idx = np.argmax(ssim_list) + 2
    max_psnr_idx = np.argmax(psnr_list) + 2

    min_ssim_idx = np.argmin(ssim_list) + 2
    min_psnr_idx = np.argmin(psnr_list) + 2

    five_max_ssim_idx = np.argsort(ssim_list)[-NBR_MARKED - 1:-1]
    five_max_psnr_idx = np.argsort(psnr_list)[-NBR_MARKED - 1:-1]

    five_min_ssim_idx = np.argsort(ssim_list)[1:NBR_MARKED + 1]
    five_min_psnr_idx = np.argsort(psnr_list)[1:NBR_MARKED + 1]

    for i in range(0, NBR_MARKED):
        ws[int(five_max_ssim_idx[i] + 2)][2].fill = color_five_best
        ws[int(five_max_psnr_idx[i] + 2)][3].fill = color_five_best

        ws[int(five_min_ssim_idx[i] + 2)][2].fill = color_five_worst
        ws[int(five_min_psnr_idx[i] + 2)][3].fill = color_five_worst

    ws[int(max_ssim_idx)][2].fill = color_best
    ws[int(max_psnr_idx)][3].fill = color_best

    ws[int(min_ssim_idx)][2].fill = color_worst
    ws[int(min_psnr_idx)][3].fill = color_worst

    for i in range(2, len(ssim_list) + 2):
        ws[i][0].value = '=HYPERLINK("{}", "im{}")'.format(ws[i][0].value, i - 1)
        ws[i][0].style = 'Hyperlink'
        ws[i][1].value = '=HYPERLINK("{}", "im{}")'.format(ws[i][1].value, i - 1)
        ws[i][1].style = 'Hyperlink'

    ws[int(len(ssim_list) + 3)][1].value = 'mean:'
    ws[int(len(ssim_list) + 4)][1].value = 'median:'
    ws[int(len(ssim_list) + 5)][1].value = 'std:'

    ws[len(ssim_list) + 3][2].value = ssim_mean
    ws[int(len(ssim_list) + 3)][3].value = psnr_mean

    ws[len(ssim_list) + 4][2].value = ssim_median
    ws[int(len(ssim_list) + 4)][3].value = psnr_median

    ws[len(ssim_list) + 5][2].value = ssim_std
    ws[int(len(ssim_list) + 5)][3].value = psnr_std

    #ws[int(len(ssim_list) + 3)][2].fill = color_mean
    #ws[int(len(ssim_list) + 3)][3].fill = color_mean
    #ws[int(len(ssim_list) + 4)][2].fill = color_median
    #ws[int(len(ssim_list) + 4)][3].fill = color_median
    #ws[int(len(ssim_list) + 5)][2].fill = color_std
    #ws[int(len(ssim_list) + 5)][3].fill = color_std

    wb.save(dir_save)


if __name__ == "__main__":

    if len(sys.argv) == 3:
        evaluate(str(sys.argv[1]), str(sys.argv[2]))
    elif len(sys.argv) == 4:
        evaluate(str(sys.argv[1]), str(sys.argv[2]), str(sys.argv[3]))
    else:
        evaluate(str(sys.argv[1]), str(sys.argv[2]), str(sys.argv[3]), str(sys.argv[4]))
