import argparse
import os
from datetime import datetime
import openpyxl
from PIL import Image
from openpyxl.styles import PatternFill
from skimage import data, img_as_float
from skimage.metrics import structural_similarity as ssim
from skimage.metrics import peak_signal_noise_ratio as psnr
import pandas as pd
import numpy as np
from tqdm import tqdm

date = str(datetime.now())
parser = argparse.ArgumentParser('Evaluation')
parser.add_argument('--background_dir', default="./background", help="path to background image folder, not used for real data")
parser.add_argument('--filtered_dir', default="./filtered", help="path to filtered image folder")
parser.add_argument('--original_data_dir', default="./original", help="path to original image folder")
parser.add_argument('--is_real_data', type=bool, default=False, help="is True if evaluating real data otherwise False")
parser.add_argument('--nbr_marked', type=int, default=5, help="nbr of best and worst metrics to mark, not used for real data")
parser.add_argument('--save_dir', default='./metrics_results_DATA_SET_NAME_{}.xlsx'.format(date),
                    help="path where to save metric results")


def is_image_file(filename):
    # check if the image ends with png
    IMG_EXTENSIONS = [
        '.jpg', '.JPG', '.jpeg', '.JPEG',
        '.png', '.PNG', '.ppm', '.PPM', '.bmp', '.BMP',
    ]
    return any(filename.endswith(extension) for extension in IMG_EXTENSIONS)


def get_col_widths(dataframe):
    # First we find the maximum length of the index column
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]


class Evaluator:
    def __init__(self, args):
        # ARGUMENTS
        self.args = args
        self.dir_bg = args.background_dir
        self.dir_res = args.filtered_dir
        self.dir_test = args.original_data_dir
        self.dir_save = args.save_dir
        if not self.dir_save.endswith('xlsx'):
            self.dir_save=os.path.normpath(self.dir_save+'/metrics_results_DATA_SET_NAME_{}.xlsx'.format(date))
        self.nbr_marked = args.nbr_marked
        self.is_real_data = args.is_real_data

        # COLORS
        self.color_mean = PatternFill(patternType='solid',
                                      fgColor='7777C9')
        self.color_median = PatternFill(patternType='solid',
                                        fgColor='4B4B9F')
        self.color_std = PatternFill(patternType='solid',
                                     fgColor='70709A')
        self.color_best = PatternFill(patternType='solid',
                                      fgColor='31A91F')
        self.color_worst = PatternFill(patternType='solid',
                                       fgColor='D81F1F')
        self.color_five_best = PatternFill(patternType='solid',
                                           fgColor='8FBC89')
        self.color_five_worst = PatternFill(patternType='solid',
                                            fgColor='DE8E7E')

    def evaluate_synthetic_data(self):
        # def evaluate(dir_bg, dir_res, dir_test, data_id='DATA SET NAME', dir_save=None):

        psnr_list = []
        ssim_list = []
        image_list_gt = []
        image_list_result = []
        image_list_test = []

        for img_bg, img_res, img_test in tqdm(zip(sorted(os.listdir(self.dir_bg), key=str.casefold), sorted(os.listdir(self.dir_res), key=str.casefold),
                                              sorted(os.listdir(self.dir_test), key=str.casefold))):
            ground_truth_image_path = os.path.join(self.dir_bg, img_bg)
            result_image_path = os.path.join(self.dir_res, img_res)
            test_image_path = os.path.join(self.dir_test, img_test)
            if is_image_file(ground_truth_image_path) and is_image_file(result_image_path) and is_image_file(
                    test_image_path):
                can_load = True
                try:
                    ground_truth_img = Image.open(ground_truth_image_path)
                    result_img = Image.open(result_image_path)
                    test_img = Image.open(test_image_path)
                except:
                    print('Was not able to load ' + ground_truth_image_path + 'or' + result_image_path)
                    can_load = False
                    continue
                if can_load:
                    image_list_gt.append(ground_truth_image_path)
                    image_list_result.append(result_image_path)
                    image_list_test.append(test_image_path)

                    ground_truth_img = img_as_float(ground_truth_img)
                    result_img = img_as_float(result_img)

                    ssim_val = ssim(ground_truth_img, result_img, multichannel=True)
                    psnr_val = psnr(ground_truth_img, result_img)
                    ssim_list.append(ssim_val)
                    psnr_list.append(psnr_val)

        d = {'GT': image_list_gt, 'Filtered': image_list_result, 'SSIM': ssim_list, 'PSNR': psnr_list,
             'Original': image_list_test}
        df = pd.DataFrame(data=d)
        df.to_excel(self.dir_save, index=False, sheet_name='sheet1')

        # mean and std values
        ssim_mean = np.mean(ssim_list)
        psnr_mean = np.mean(psnr_list)
        ssim_median = np.median(ssim_list)
        psnr_median = np.median(psnr_list)
        ssim_std = np.std(ssim_list)
        psnr_std = np.std(psnr_list)

        wb = openpyxl.load_workbook(self.dir_save)
        ws = wb['sheet1']  # Name of the working sheet

        max_ssim_idx = np.argmax(ssim_list) + 2
        max_psnr_idx = np.argmax(psnr_list) + 2

        min_ssim_idx = np.argmin(ssim_list) + 2
        min_psnr_idx = np.argmin(psnr_list) + 2

        five_max_ssim_idx = np.argsort(ssim_list)[-self.nbr_marked - 1:-1]
        five_max_psnr_idx = np.argsort(psnr_list)[-self.nbr_marked - 1:-1]

        five_min_ssim_idx = np.argsort(ssim_list)[1:self.nbr_marked + 1]
        five_min_psnr_idx = np.argsort(psnr_list)[1:self.nbr_marked + 1]

        for i in range(0, self.nbr_marked):
            ws[int(five_max_ssim_idx[i] + 2)][2].fill = self.color_five_best
            ws[int(five_max_psnr_idx[i] + 2)][3].fill = self.color_five_best

            ws[int(five_min_ssim_idx[i] + 2)][2].fill = self.color_five_worst
            ws[int(five_min_psnr_idx[i] + 2)][3].fill = self.color_five_worst

        ws[int(max_ssim_idx)][2].fill = self.color_best
        ws[int(max_psnr_idx)][3].fill = self.color_best

        ws[int(min_ssim_idx)][2].fill = self.color_worst
        ws[int(min_psnr_idx)][3].fill = self.color_worst

        for i in range(2, len(ssim_list) + 2):
            ws[i][0].value = '=HYPERLINK("{}", "im{}")'.format(ws[i][0].value, i - 1)
            ws[i][0].style = 'Hyperlink'
            ws[i][1].value = '=HYPERLINK("{}", "im{}")'.format(ws[i][1].value, i - 1)
            ws[i][1].style = 'Hyperlink'
            ws[i][4].value = '=HYPERLINK("{}", "im{}")'.format(ws[i][4].value, i - 1)
            ws[i][4].style = 'Hyperlink'

        ws[int(len(ssim_list) + 3)][1].value = 'mean:'
        ws[int(len(ssim_list) + 4)][1].value = 'median:'
        ws[int(len(ssim_list) + 5)][1].value = 'std:'

        ws[len(ssim_list) + 3][2].value = ssim_mean
        ws[int(len(ssim_list) + 3)][3].value = psnr_mean

        ws[len(ssim_list) + 4][2].value = ssim_median
        ws[int(len(ssim_list) + 4)][3].value = psnr_median

        ws[len(ssim_list) + 5][2].value = ssim_std
        ws[int(len(ssim_list) + 5)][3].value = psnr_std

        wb.save(self.dir_save)
        print("results saved in: " + self.dir_save)

    def evaluate_real_data(self):
        rank_list = []
        image_list_original = []
        image_list_result = []

        for img_test, img_res in tqdm(zip(sorted(os.listdir(self.dir_test),key=str.casefold), sorted(os.listdir(self.dir_res), key=str.casefold))):
            original_image_path = os.path.join(self.dir_test, img_test)
            result_image_path = os.path.join(self.dir_res, img_res)
            if is_image_file(original_image_path) and is_image_file(result_image_path):
                can_load = True
                try:
                    ground_truth_img = Image.open(original_image_path)
                    result_img = Image.open(result_image_path)
                except:
                    print('Was not able to load ' + original_image_path + 'or' + result_image_path)
                    can_load = False
                    continue
                if can_load:
                    image_list_original.append(original_image_path)
                    image_list_result.append(result_image_path)
                    rank_list.append(0)

        d = {'Original': image_list_original, 'Filtered': image_list_result, 'Rank': rank_list}
        df = pd.DataFrame(data=d)
        df.to_excel(self.dir_save, index=False, sheet_name='sheet1')

        wb = openpyxl.load_workbook(self.dir_save)
        ws = wb['sheet1']  # Name of the working sheet

        for i in range(2, len(rank_list) + 2):
            ws[i][0].value = '=HYPERLINK("{}", "im{}")'.format(ws[i][0].value, i - 1)
            ws[i][0].style = 'Hyperlink'
            ws[i][1].value = '=HYPERLINK("{}", "im{}")'.format(ws[i][1].value, i - 1)
            ws[i][1].style = 'Hyperlink'

        ws[int(len(rank_list) + 3)][1].value = 'mean:'
        ws[int(len(rank_list) + 4)][1].value = 'median:'
        ws[int(len(rank_list) + 5)][1].value = 'std:'

        ws[len(rank_list) + 3][2].value = '=AVERAGE(C{}:C{})'.format(2, len(rank_list) + 1)
        ws[len(rank_list) + 4][2].value = '=MEDIAN(C{}:C{})'.format(2, len(rank_list) + 1)
        ws[len(rank_list) + 5][2].value = '=STDEV(C{}:C{})'.format(2, len(rank_list) + 1)

        wb.save(self.dir_save)
        print("results saved in: " + self.dir_save)


def main():
    args = parser.parse_args()
    evaluator = Evaluator(args)
    if not args.is_real_data:
        evaluator.evaluate_synthetic_data()
    else:
        evaluator.evaluate_real_data()


if __name__ == "__main__":
    main()
